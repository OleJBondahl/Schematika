"""Tests for Project BOM aggregation and export internals.

Targets the survivors clustered in:
- ``Project._aggregate_bom`` (devices, terminals, PLC modules)
- ``Project._export_bom_csv``
- ``Project._export_bom_excel``
- ``Project._generate_bom_typst``
- ``Project._count_terminal_pins``
"""

from __future__ import annotations

import csv
import os

from openpyxl import load_workbook

from schematika.electrical import Terminal
from schematika.electrical.builder import BuildResult
from schematika.electrical.internal_device import InternalDevice
from schematika.electrical.plc_resolver import PlcModuleType
from schematika.electrical.system.system import Circuit
from schematika.project import Project

# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------


def _project_with_devices_and_terminals() -> Project:
    """Build a small Project with 3 circuits worth of device + terminal data."""
    contactor = InternalDevice(prefix="K", mpn="ABB-AF09", description="Contactor 9A")
    fuse = InternalDevice(prefix="F", mpn="SIEM-5SY", description="MCB 6A")

    def builder_motors(state, **_kw):
        return BuildResult(
            state=state,
            circuit=Circuit(),
            used_terminals=[],
            device_registry={"K1": contactor, "K2": contactor, "K10": contactor},
        )

    def builder_protection(state, **_kw):
        return BuildResult(
            state=state,
            circuit=Circuit(),
            used_terminals=[],
            device_registry={"F1": fuse, "F2": fuse},
        )

    def builder_empty(state, **_kw):
        return BuildResult(state=state, circuit=Circuit(), used_terminals=[])

    p = Project()
    p.terminals(
        Terminal("X1", title="Power", description="Phoenix", mpn="3209510"),
        Terminal("X2", title="GND", description="Phoenix", mpn="3209510"),
        Terminal("X3", title="Signal", description="Wago", mpn="2002-1201"),
        Terminal("PLC:DO", title="DO ref", reference=True, mpn="ignored"),
    )
    p.add_circuit("motors", builder_motors)
    p.add_circuit("protection", builder_protection)
    p.add_circuit("empty", builder_empty)
    p.build_circuits()
    return p


# ---------------------------------------------------------------------------
# _aggregate_bom — device grouping
# ---------------------------------------------------------------------------


def test_aggregate_bom_groups_devices_by_mpn():
    p = _project_with_devices_and_terminals()
    rows = p._aggregate_bom()
    contactor_rows = [r for r in rows if r[1] == "ABB-AF09"]
    assert len(contactor_rows) == 1
    _tags, mpn, desc, qty = contactor_rows[0]
    assert mpn == "ABB-AF09"
    assert desc == "Contactor 9A"
    assert qty == 3


def test_aggregate_bom_device_tags_natural_sorted():
    p = _project_with_devices_and_terminals()
    rows = p._aggregate_bom()
    contactor_row = next(r for r in rows if r[1] == "ABB-AF09")
    # K10 should come after K2 with natural sort
    assert contactor_row[0] == "K1/K2/K10"


def test_aggregate_bom_separate_devices_separate_rows():
    p = _project_with_devices_and_terminals()
    rows = p._aggregate_bom()
    mpns = [r[1] for r in rows]
    assert "ABB-AF09" in mpns
    assert "SIEM-5SY" in mpns


def test_aggregate_bom_qty_matches_unique_tags():
    p = _project_with_devices_and_terminals()
    rows = p._aggregate_bom()
    fuse_row = next(r for r in rows if r[1] == "SIEM-5SY")
    assert fuse_row[3] == 2  # F1, F2


def test_aggregate_bom_dedupes_repeated_tags():
    """If a tag appears twice in a registry, BOM should still count it once."""
    contactor = InternalDevice(prefix="K", mpn="ABB-AF09", description="Contactor 9A")

    def builder1(state, **_kw):
        return BuildResult(
            state=state,
            circuit=Circuit(),
            used_terminals=[],
            device_registry={"K1": contactor},
        )

    def builder2(state, **_kw):
        return BuildResult(
            state=state,
            circuit=Circuit(),
            used_terminals=[],
            device_registry={"K1": contactor},  # same tag in a second circuit
        )

    p = Project()
    p.add_circuit("c1", builder1)
    p.add_circuit("c2", builder2)
    p.build_circuits()
    rows = p._aggregate_bom()
    contactor_row = next(r for r in rows if r[1] == "ABB-AF09")
    assert contactor_row[3] == 1
    assert contactor_row[0] == "K1"


# ---------------------------------------------------------------------------
# _aggregate_bom — terminal grouping
# ---------------------------------------------------------------------------


def test_aggregate_bom_groups_terminals_by_mpn():
    p = _project_with_devices_and_terminals()
    rows = p._aggregate_bom()
    phoenix_rows = [r for r in rows if r[1] == "3209510"]
    assert len(phoenix_rows) == 1
    tags, _, desc, _ = phoenix_rows[0]
    assert "X1" in tags
    assert "X2" in tags
    assert desc == "Phoenix"


def test_aggregate_bom_excludes_reference_terminals():
    p = _project_with_devices_and_terminals()
    rows = p._aggregate_bom()
    # PLC:DO is a reference terminal -- must not be in BOM
    for tags, _mpn, _desc, _qty in rows:
        assert "PLC:DO" not in tags


def test_aggregate_bom_terminal_without_mpn_excluded():
    p = Project()
    p.terminals(Terminal("X1", title="No MPN"))
    p.build_circuits()
    rows = p._aggregate_bom()
    assert rows == []


def test_aggregate_bom_terminal_qty_counts_pins():
    """Terminal qty in the BOM equals total used pin count across that MPN."""
    t1 = Terminal("X1", description="Phoenix", mpn="3209510")
    p = Project()
    p.terminals(t1)
    # Add an external connection that uses pin 1 of X1
    p.external_connections([("DEV1", "out", "X1", "1", "PLC:DI", "")])
    p.external_connections([("DEV2", "out", "X1", "2", "PLC:DI", "")])
    p.build_circuits()
    rows = p._aggregate_bom()
    phx_row = next(r for r in rows if r[1] == "3209510")
    assert phx_row[3] == 2  # 2 pins used


# ---------------------------------------------------------------------------
# _aggregate_bom — PLC modules
# ---------------------------------------------------------------------------


def test_aggregate_bom_includes_plc_modules():
    do_module = PlcModuleType(
        mpn="750-1504", signal_type="DO", channels=4, pins_per_channel=("",)
    )
    rack = [("DO1", do_module), ("DO2", do_module)]
    p = Project()
    p.plc_rack(rack)
    p.build_circuits()
    rows = p._aggregate_bom()
    plc_rows = [r for r in rows if r[1] == "750-1504"]
    assert len(plc_rows) == 1
    tags, _mpn, desc, qty = plc_rows[0]
    assert tags == "DO1/DO2"
    assert desc == "PLC DO module (4ch)"
    assert qty == 2


def test_aggregate_bom_no_plc_when_rack_unset():
    p = Project()
    p.build_circuits()
    rows = p._aggregate_bom()
    # No rack -> no PLC rows
    assert rows == []


def test_aggregate_bom_plc_groups_distinct_mpns():
    do_module = PlcModuleType(
        mpn="750-1504", signal_type="DO", channels=4, pins_per_channel=("",)
    )
    di_module = PlcModuleType(
        mpn="750-1405", signal_type="DI", channels=8, pins_per_channel=("",)
    )
    rack = [("DI1", di_module), ("DO1", do_module), ("DI2", di_module)]
    p = Project()
    p.plc_rack(rack)
    p.build_circuits()
    rows = p._aggregate_bom()
    di_row = next(r for r in rows if r[1] == "750-1405")
    do_row = next(r for r in rows if r[1] == "750-1504")
    assert di_row[3] == 2  # 2 DI modules
    assert do_row[3] == 1
    assert "(8ch)" in di_row[2]


# ---------------------------------------------------------------------------
# _aggregate_bom — section ordering: devices, terminals, PLC
# ---------------------------------------------------------------------------


def test_aggregate_bom_section_order_devices_terminals_plc():
    do_module = PlcModuleType(
        mpn="750-1504", signal_type="DO", channels=4, pins_per_channel=("",)
    )
    p = _project_with_devices_and_terminals()
    p.plc_rack([("DO1", do_module)])
    rows = p._aggregate_bom()

    def section_for(row):
        mpn = row[1]
        if mpn in ("ABB-AF09", "SIEM-5SY"):
            return 0  # device
        if mpn in ("3209510", "2002-1201"):
            return 1  # terminal
        return 2  # plc

    sections = [section_for(r) for r in rows]
    assert sections == sorted(sections)


# ---------------------------------------------------------------------------
# _export_bom_csv
# ---------------------------------------------------------------------------


def test_export_bom_csv_writes_header_and_rows(tmp_path):
    p = _project_with_devices_and_terminals()
    csv_path = str(tmp_path / "bom.csv")
    p.export_bom_csv(csv_path)
    p._export_bom_csv()

    assert os.path.exists(csv_path)
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    assert rows[0] == ["Tags", "MPN", "Description", "Qty"]
    # At least one device row + one terminal row
    assert len(rows) >= 3


def test_export_bom_csv_no_op_when_unset(tmp_path):
    """_export_bom_csv must do nothing when no path was registered."""
    p = _project_with_devices_and_terminals()
    p._export_bom_csv()  # no path set -> early return
    # No file should be created in tmp_path
    assert os.listdir(str(tmp_path)) == []


def test_export_bom_csv_qty_column_is_int(tmp_path):
    p = _project_with_devices_and_terminals()
    csv_path = str(tmp_path / "bom.csv")
    p.export_bom_csv(csv_path)
    p._export_bom_csv()

    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)  # header
        for row in reader:
            # Qty is always a non-negative integer
            assert row[3].isdigit()


def test_export_bom_csv_creates_parent_dir(tmp_path):
    p = _project_with_devices_and_terminals()
    csv_path = str(tmp_path / "deep" / "nested" / "bom.csv")
    p.export_bom_csv(csv_path)
    p._export_bom_csv()
    assert os.path.exists(csv_path)


def test_export_bom_csv_returns_self():
    """export_bom_csv() should return self for chaining."""
    p = Project()
    result = p.export_bom_csv("/tmp/bom.csv")
    assert result is p


# ---------------------------------------------------------------------------
# _export_bom_excel
# ---------------------------------------------------------------------------


def test_export_bom_excel_creates_workbook(tmp_path):
    p = _project_with_devices_and_terminals()
    xlsx_path = str(tmp_path / "bom.xlsx")
    p.export_bom_excel(xlsx_path)
    p._export_bom_excel()
    assert os.path.exists(xlsx_path)


def test_export_bom_excel_header_row_is_styled(tmp_path):
    p = _project_with_devices_and_terminals()
    xlsx_path = str(tmp_path / "bom.xlsx")
    p.export_bom_excel(xlsx_path)
    p._export_bom_excel()

    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws.title == "BOM"
    # Header values
    assert ws.cell(row=1, column=1).value == "Tags"
    assert ws.cell(row=1, column=2).value == "MPN"
    assert ws.cell(row=1, column=3).value == "Description"
    assert ws.cell(row=1, column=4).value == "Qty"
    # Bold + grey fill on header
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=1, column=1).alignment.horizontal == "left"


def test_export_bom_excel_data_rows_match_aggregate(tmp_path):
    p = _project_with_devices_and_terminals()
    xlsx_path = str(tmp_path / "bom.xlsx")
    p.export_bom_excel(xlsx_path)
    p._export_bom_excel()

    expected = p._aggregate_bom()
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for idx, (tags, mpn, desc, qty) in enumerate(expected, start=2):
        assert ws.cell(row=idx, column=1).value == tags
        assert ws.cell(row=idx, column=2).value == mpn
        assert ws.cell(row=idx, column=3).value == desc
        assert ws.cell(row=idx, column=4).value == qty


def test_export_bom_excel_qty_aligned_right(tmp_path):
    p = _project_with_devices_and_terminals()
    xlsx_path = str(tmp_path / "bom.xlsx")
    p.export_bom_excel(xlsx_path)
    p._export_bom_excel()

    wb = load_workbook(xlsx_path)
    ws = wb.active
    # First data row qty cell should be right-aligned
    assert ws.cell(row=2, column=4).alignment.horizontal == "right"


def test_export_bom_excel_column_widths(tmp_path):
    p = _project_with_devices_and_terminals()
    xlsx_path = str(tmp_path / "bom.xlsx")
    p.export_bom_excel(xlsx_path)
    p._export_bom_excel()

    wb = load_workbook(xlsx_path)
    ws = wb.active
    # Column widths set in _export_bom_excel: A=30, B=20, C=40, D=8
    assert ws.column_dimensions["A"].width == 30
    assert ws.column_dimensions["B"].width == 20
    assert ws.column_dimensions["C"].width == 40
    assert ws.column_dimensions["D"].width == 8


def test_export_bom_excel_no_op_when_unset(tmp_path):
    p = _project_with_devices_and_terminals()
    p._export_bom_excel()  # no path registered
    assert os.listdir(str(tmp_path)) == []


def test_export_bom_excel_creates_parent_dir(tmp_path):
    p = _project_with_devices_and_terminals()
    xlsx_path = str(tmp_path / "deep" / "bom.xlsx")
    p.export_bom_excel(xlsx_path)
    p._export_bom_excel()
    assert os.path.exists(xlsx_path)


def test_export_bom_excel_returns_self():
    p = Project()
    result = p.export_bom_excel("/tmp/bom.xlsx")
    assert result is p


# ---------------------------------------------------------------------------
# _generate_bom_typst
# ---------------------------------------------------------------------------


def test_generate_bom_typst_contains_title():
    p = Project()
    out = p._generate_bom_typst([])
    assert "Bill of Materials" in out


def test_generate_bom_typst_has_table_header():
    p = Project()
    out = p._generate_bom_typst([])
    assert "[Tags]" in out
    assert "[MPN]" in out
    assert "[Description]" in out
    assert "[Qty]" in out


def test_generate_bom_typst_emits_one_row_per_entry():
    p = Project()
    rows = [
        ("Q1/Q2", "ABB-AF09", "Contactor", 2),
        ("F1", "SIEM-5SY", "MCB", 1),
    ]
    out = p._generate_bom_typst(rows)
    assert "Q1/Q2" in out
    assert "ABB-AF09" in out
    assert "Contactor" in out
    assert "SIEM-5SY" in out
    # Qty values should appear
    assert "[2]" in out
    assert "[1]" in out


def test_generate_bom_typst_escapes_hash():
    """Hash characters are special in Typst markup and must be escaped."""
    p = Project()
    rows = [("T#1", "MPN#X", "Desc#Y", 1)]
    out = p._generate_bom_typst(rows)
    assert "T\\#1" in out
    assert "MPN\\#X" in out
    assert "Desc\\#Y" in out
    assert "T#1" not in out.replace("T\\#1", "")


def test_generate_bom_typst_columns_layout():
    """Output must declare 2-column layout & 4-column table widths."""
    p = Project()
    out = p._generate_bom_typst([])
    assert "#columns(2" in out
    assert "columns: (4.5cm, 3.5cm, 1fr, 1cm)" in out


def test_generate_bom_typst_empty_rows():
    p = Project()
    out = p._generate_bom_typst([])
    # Still produces structurally valid markup
    assert "#table" in out
    assert "Bill of Materials" in out


# ---------------------------------------------------------------------------
# _count_terminal_pins
# ---------------------------------------------------------------------------


def test_count_terminal_pins_external_connections():
    p = Project()
    p.external_connections(
        [
            ("DEV", "p", "X1", "1", "PLC:DI", ""),
            ("DEV", "p", "X1", "2", "PLC:DI", ""),
            ("DEV", "p", "X1", "1", "PLC:DI", ""),  # duplicate pin -> dedup
            ("DEV", "p", "X2", "5", "PLC:DI", ""),
        ]
    )
    counts = p._count_terminal_pins()
    assert counts["X1"] == 2  # pins 1, 2 (dedup)
    assert counts["X2"] == 1


def test_count_terminal_pins_skips_empty_rows():
    p = Project()
    # Empty pin must be skipped
    p.external_connections([("DEV", "p", "X1", "", "PLC:DI", "")])
    p.external_connections([("DEV", "p", "", "1", "PLC:DI", "")])
    counts = p._count_terminal_pins()
    assert counts == {}


def test_count_terminal_pins_includes_terminal_only_connections():
    p = Project()
    p.internal_wiring([("DEV", "p", "X1", "1", "DEV2", "p2")])
    p.internal_wiring([("DEV", "p", "X1", "2", "DEV2", "p2")])
    counts = p._count_terminal_pins()
    assert counts["X1"] == 2
