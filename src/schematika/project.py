"""Project class -- Layer 0 declarative API for Schematika."""

import os
import shutil
from collections.abc import Callable, Iterable
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import TYPE_CHECKING, Any

from schematika.core.exceptions import CircuitValidationError
from schematika.electrical.builder import BuildResult, CircuitBuilder
from schematika.electrical.builder_models import BridgeMode
from schematika.electrical.harness import Harness, HarnessBuildResult, PlcAssignment

if TYPE_CHECKING:
    from schematika.catalog.cables import CableRegistry
    from schematika.catalog.identifiers import NetId
    from schematika.catalog.refs import PinRef
    from schematika.catalog.registry import DeviceCatalog
    from schematika.catalog.wires import Wire
    from schematika.electrical.field_devices import ConnectionRow
    from schematika.electrical.harness import Plc
    from schematika.electrical.model.state import GenerationState
    from schematika.electrical.plc_resolver import PlcRack
    from schematika.electrical.terminal_sidecar import TerminalWireFact
    from schematika.pcb.model import PCBBuildResult
    from schematika.pid.builder import PIDBuildResult
    from schematika.rendering.typst.compiler import TypstCompiler

from schematika.core.options import DescriptorBuildOptions
from schematika.electrical.descriptors import Descriptor, build_from_descriptors
from schematika.electrical.system.connection_registry import (
    export_registry_to_csv,
    get_registry,
)
from schematika.electrical.system.system import render_system
from schematika.electrical.terminal import Terminal
from schematika.electrical.utils.autonumbering import create_autonumberer
from schematika.electrical.utils.export_utils import (
    export_terminal_list,
    finalize_terminal_csv,
)

# ---------------------------------------------------------------------------
# Internal data structures
# ---------------------------------------------------------------------------


@dataclass
class _CircuitDef:
    """Internal deferred circuit definition."""

    key: str
    factory: str  # "descriptors" or "custom"
    params: dict[str, Any] = field(default_factory=dict)
    count: int = 1
    wire_labels: list[str] | None = None
    reuse_tags: dict[str, str] | None = None  # maps prefix -> circuit key
    components: list[Descriptor] | None = None
    builder_fn: Callable | None = None
    start_indices: dict[str, int] | None = None
    terminal_start_indices: dict[str, int] | None = None


@dataclass
class _PageDef:
    """Internal page definition."""

    # page_type values: "schematic", "front", "terminal_report", "plc_report",
    # "custom", "bom_report", "pid", "cable", "cable_toc"
    page_type: str
    title: str = ""
    circuit_key: str = ""
    circuit_keys: list[str] | None = None
    md_path: str = ""
    notice: str | None = None
    csv_path: str = ""
    typst_content: str = ""
    cable_entries: list[tuple[str, str, str, str]] | None = None
    cable_toc_entries: list[tuple[str, str, str]] | None = None


@dataclass
class _PIDDef:
    """Internal deferred P&ID diagram definition."""

    key: str
    builder_or_factory: Any  # PIDBuilder instance or callable(state) -> PIDBuildResult


def _resolve_svg_for_page(
    page_type: str,
    key: str,
    svg_paths: dict[str, str],
    csv_paths: dict[str, str],
    pid_svg_paths: dict[str, str] | None,
) -> tuple[str, str | None]:
    """Resolve SVG and CSV paths for a schematic/pid page."""
    if page_type == "pid":
        return (pid_svg_paths or {}).get(key, ""), None
    return svg_paths.get(key, ""), csv_paths.get(key)


def _render_with_optional_pcb_viewbox(
    circuit: Any,  # noqa: ANN401
    svg_path: str,
    pcb_dims: tuple[str, float, float] | None,
) -> None:
    """Render a circuit; if pcb_dims is set, force width/height/viewBox to match."""
    if pcb_dims is None:
        render_system(circuit, svg_path)
    else:
        viewbox, width, height = pcb_dims
        render_system(
            circuit, svg_path, width=int(width), height=int(height), viewbox=viewbox
        )


def _wires_to_terminal_rows(result: HarnessBuildResult) -> "list[ConnectionRow]":
    """Convert harness wires into terminal-only ConnectionRow tuples.

    Each 2-point wire ``source -> target`` becomes
    ``(component, pin, terminal, terminal_pin, "", "")`` — the same shape
    ``internal_wiring()`` accepts. The Plc-waypoint case is out of C1a scope
    (field devices stay on the legacy path).
    """
    return [
        (
            str(wire.source.device),
            wire.source.port_id,
            str(wire.target.device),
            wire.target.port_id,
            "",
            "",
        )
        for wire in result.wires
    ]


def _wires_to_terminal_facts(
    result: HarnessBuildResult,
) -> "list[tuple[Wire, TerminalWireFact]]":
    """Pair each harness wire with the first-waypoint-anchors terminal fact.

    Each 2-point wire ``source -> target`` becomes a ``(wire, fact)`` pair with
    ``anchor="source"`` (first waypoint is the key terminal) and ``side="bottom"``
    (the other endpoint goes to the TO columns) -- byte-identical to the
    hand-built ``internal_wiring`` tuple ``("", "", source, target)``.
    """
    from schematika.electrical.terminal_sidecar import TerminalWireFact

    return [
        (wire, TerminalWireFact(anchor="source", side="bottom"))
        for wire in result.wires
    ]


# ---------------------------------------------------------------------------
# Project class
# ---------------------------------------------------------------------------


class Project:
    """Top-level mutable builder that assembles a complete multi-page schematic.

    Threads autonumbering state across all :meth:`add_page` calls and
    collects circuit pages, title-block metadata, and BOM data.
    Call :meth:`build` once to produce the final SVG output; do not reuse
    a ``Project`` instance after :meth:`build`.

    Examples:
        >>> from schematika import Project
        >>> p = Project(title="Test Panel", drawing_number="D-001")
        >>> p.title
        'Test Panel'
    """

    def __init__(  # noqa: PLR0913
        self,
        title: str = "",
        drawing_number: str = "",
        author: str = "",
        project: str = "",
        revision: str = "00",
        logo: str | None = None,
        font: str = "Times New Roman",
        *,
        sort_integer_pins: bool = True,
        sort_alphabetic_pins: bool = False,
    ) -> None:
        """Stores title-block metadata; no I/O until `build()`."""
        from schematika.cable.builder import PinSortConfig

        self.title = title
        self.drawing_number = drawing_number
        self.author = author
        self.project = project
        self.revision = revision
        self.logo = logo
        self.font = font
        self.sort_integer_pins = sort_integer_pins
        self.sort_alphabetic_pins = sort_alphabetic_pins
        self._pin_sort = PinSortConfig(
            sort_integers=sort_integer_pins, sort_alphabetic=sort_alphabetic_pins
        )

        self._state = create_autonumberer()
        self._terminals: dict[str, Terminal] = {}
        self._circuit_defs: list[_CircuitDef] = []
        self._pages: list[_PageDef] = []
        self._results: dict[str, BuildResult] = {}
        self._plc_rack: PlcRack | None = None
        self._external_connections: list[ConnectionRow] = []
        self._terminal_only_connections: list[ConnectionRow] = []
        self._route_terminal_rows: list[ConnectionRow] = []
        self._field_device_rows: list[ConnectionRow] = []
        self._field_device_wire_rows: list[list[str]] = []
        self._field_device_wires: list[Wire] = []
        self._plc_assignments: list[PlcAssignment] = []
        self._native_terminal_emit: bool = False
        self._native_plc_report: bool = False
        self._field_device_defs: list[tuple[list, dict | None, dict | None]] = []
        self._cable_runs: list = []
        self._route_decls: list[tuple[tuple[PinRef | Plc, ...], NetId | None]] = []
        self._added_wires: list[Wire] = []
        self._wire_label_export: tuple[str, dict[str, str] | None] | None = None
        self._taglist_export: str | None = None
        self._bom_excel_export: str | None = None
        self._bom_csv_export: str | None = None
        self._catalog: DeviceCatalog | None = None
        self._cable_registry: CableRegistry | None = None
        self._pid_defs: list[_PIDDef] = []
        self._pid_results: dict[str, PIDBuildResult] = {}
        self._pcb_page_viewboxes: dict[str, tuple[str, float, float]] = {}
        # circuit_key -> (viewBox_str, width_mm, height_mm)

    # ------------------------------------------------------------------
    # Terminal registration
    # ------------------------------------------------------------------

    def terminals(self, *terminals: Terminal) -> None:
        """Must be called before registering circuits that reference these terminals."""
        for t in terminals:
            self._terminals[str(t)] = t

    @property
    def _terminal_descriptions(self) -> dict[str, str]:
        return {tag: t.title for tag, t in self._terminals.items() if t.title}

    def set_pin_start(self, terminal_id: str, pin: int) -> None:
        """Next auto-allocation will be `pin + 1`; raises any per-prefix floor too."""
        tag_key = str(terminal_id)

        new_counters = {**self._state.terminal_counters, tag_key: pin}

        prefix_counters = self._state.terminal_prefix_counters
        if tag_key in prefix_counters:
            new_tag_prefixes = prefix_counters[tag_key].copy()
            for p in new_tag_prefixes:
                new_tag_prefixes[p] = pin
            new_prefix_counters = {**prefix_counters, tag_key: new_tag_prefixes}
        else:
            new_prefix_counters = prefix_counters

        self._state = replace(
            self._state,
            terminal_counters=new_counters,
            terminal_prefix_counters=new_prefix_counters,
        )

    # ------------------------------------------------------------------
    # Circuit registration
    # ------------------------------------------------------------------

    def circuit(
        self,
        key: str,
        components: list[Descriptor],
        count: int = 1,
        wire_labels: list[str] | None = None,
        reuse_tags: dict[str, str] | None = None,
        start_indices: dict[str, int] | None = None,
        terminal_start_indices: dict[str, int] | None = None,
        **kwargs: Any,  # noqa: ANN401
    ) -> None:
        """Register a custom inline circuit from descriptors."""
        self._circuit_defs.append(
            _CircuitDef(
                key=key,
                factory="descriptors",
                count=count,
                wire_labels=wire_labels,
                reuse_tags=reuse_tags,
                components=components,
                params=kwargs,
                start_indices=start_indices,
                terminal_start_indices=terminal_start_indices,
            )
        )

    def add_circuit(
        self,
        key: str,
        builder_fn: Callable,
        count: int = 1,
        **kwargs: Any,  # noqa: ANN401
    ) -> None:
        """`builder_fn(state, **kwargs)` -> BuildResult or `(state, circuit, terms)`."""
        self._circuit_defs.append(
            _CircuitDef(
                key=key,
                factory="custom",
                count=count,
                builder_fn=builder_fn,
                params=kwargs,
            )
        )

    def reserve_pins(self, key: str, terminal: "Terminal", count: int) -> "Project":
        """Advances the pin counter and registers a bridge group; circuit is empty."""

        def _reserve_fn(state: "GenerationState", **_kwargs: Any) -> BuildResult:  # noqa: ANN401
            from schematika.electrical.system.system import Circuit
            from schematika.electrical.utils.autonumbering import (
                get_terminal_counter,
                set_terminal_counter,
            )

            tag = str(terminal)
            start = get_terminal_counter(state, tag) + 1
            end = start + count - 1
            state = set_terminal_counter(state, terminal, end)
            return BuildResult(
                state=state,
                circuit=Circuit(),
                used_terminals=[],
                bridge_groups={tag: [(start, end)]},
            )

        self._circuit_defs.append(
            _CircuitDef(key=key, factory="custom", builder_fn=_reserve_fn)
        )
        return self

    # ------------------------------------------------------------------
    # Catalog and P&ID registration
    # ------------------------------------------------------------------

    def set_catalog(self, catalog: "DeviceCatalog") -> "Project":
        """Store a device catalog for cross-referencing between P&ID and electrical."""
        self._catalog = catalog
        return self

    @property
    def catalog(self) -> "DeviceCatalog | None":
        """The registered device catalog, or None if not set."""
        return self._catalog

    def add_pid(self, key: str, builder_or_factory: Any) -> "Project":  # noqa: ANN401
        """Built during `build()`; accepts PIDBuilder or `(state) -> PIDBuildResult`."""
        self._pid_defs.append(_PIDDef(key=key, builder_or_factory=builder_or_factory))
        return self

    def pid_page(self, title: str, diagram_key: str) -> "Project":
        """*diagram_key* must have been registered via `add_pid()`."""
        self._pages.append(
            _PageDef(page_type="pid", title=title, circuit_key=diagram_key)
        )
        return self

    # ------------------------------------------------------------------
    # Cable registry
    # ------------------------------------------------------------------

    def set_cable_registry(self, registry: "CableRegistry") -> "Project":
        """Store a cable registry for cross-referencing across all modules."""
        self._cable_registry = registry
        return self

    @property
    def cable_registry(self) -> "CableRegistry | None":
        """The registered cable registry, or None if not set."""
        return self._cable_registry

    # ------------------------------------------------------------------
    # Block diagram registration
    # ------------------------------------------------------------------

    def add_pcb(self, result: "PCBBuildResult") -> "Project":
        """Render PCB connector blocks and floating parts as schematic Circuits."""
        from schematika.core.state import create_initial_state
        from schematika.electrical.builder import BuildResult
        from schematika.electrical.system.system import Circuit, merge_circuits
        from schematika.pcb.render import render_connector_block, render_floating_part

        if result.state is not None:
            self._state = result.state
        else:
            self._state = create_initial_state()

        if not hasattr(self, "_pcb_page_viewboxes"):
            self._pcb_page_viewboxes = {}

        block_by_ref = {b.connector_ref: b for b in result.connector_blocks}
        mapping = result.mapping
        ir = result.ir

        def _circuit_fn(c: Any) -> Any:  # noqa: ANN401
            return lambda state, **_kw: BuildResult(
                state=state, circuit=c, used_terminals=[]
            )

        page_w, page_h = result.page_size
        page_viewbox = f"0 0 {page_w} {page_h}"
        page_dims = (page_viewbox, page_w, page_h)

        for page in result.pages:
            page_keys: list[str] = []

            # Merge connector blocks and inline floating slices into one circuit
            # per page so they share the same SVG / Typst page.
            if page.placements or page.floating_placements:
                from schematika.pcb.model import FloatingPart

                merged = Circuit()
                for block_ref, origin_x, origin_y in page.placements:
                    block = block_by_ref[block_ref]
                    block_circuit = render_connector_block(
                        block,
                        mapping,
                        origin_x_mm=origin_x,
                        origin_y_mm=origin_y,
                        layout=result.layout,
                    )
                    merged = merge_circuits(merged, block_circuit)
                for fsp in page.floating_placements:
                    single_slice = FloatingPart(
                        part_ref=fsp.part_ref,
                        slice_indices=(fsp.slice_index,),
                    )
                    floating_circuit = render_floating_part(
                        single_slice,
                        mapping=mapping,
                        ir=ir,
                        origin_x_mm=fsp.x_mm,
                        origin_y_mm=fsp.y_mm,
                        layout=result.layout,
                    )
                    merged = merge_circuits(merged, floating_circuit)

                page_key = f"pcb_page_{page.title}"
                self.add_circuit(page_key, _circuit_fn(merged))
                # Record the page-extent viewBox + dimensions so render_system uses
                # page coords with matching width/height instead of content-fitted.
                self._pcb_page_viewboxes[page_key] = page_dims
                page_keys.append(page_key)

            if page_keys:
                self.page(page.title, page_keys)
        return self

    # ------------------------------------------------------------------
    # PLC rack and external connections
    # ------------------------------------------------------------------

    def plc_rack(self, rack: "PlcRack") -> "Project":
        """Once registered, `build()` auto-generates the PLC connections CSV."""
        self._plc_rack = rack
        return self

    def use_native_terminal_emit(self) -> "Project":
        """Opt-in: emit the terminal CSV via the native Wire+sidecar path (C3a).

        Examples:
            >>> from schematika.project import Project
            >>> isinstance(Project().use_native_terminal_emit(), Project)
            True
        """
        self._native_terminal_emit = True
        return self

    def use_native_plc_report(self) -> "Project":
        """Opt-in: emit the PLC report via the native HarnessBuildResult path (R3).

        Examples:
            >>> from schematika.project import Project
            >>> isinstance(Project().use_native_plc_report(), Project)
            True
        """
        self._native_plc_report = True
        return self

    def external_connections(self, connections: "list[ConnectionRow]") -> "Project":
        """Field-to-cabinet connections; resolved against the PLC rack."""
        self._external_connections.extend(connections)
        return self

    def internal_wiring(self, connections: "list[ConnectionRow]") -> "Project":
        """Terminal-to-terminal connections; reported but not in cable exports."""
        self._terminal_only_connections.extend(connections)
        return self

    def add_field_devices(
        self,
        connections: "list[ConnectionRow]",
        reuse_terminals: dict[str, str] | None = None,  # noqa: ARG002
    ) -> "Project":
        """Alias for `external_connections()`; `reuse_terminals` reserved (ignored)."""
        self._external_connections.extend(connections)
        return self

    def field_devices(
        self,
        devices: list,
        reuse_terminals: dict | None = None,
        template_reuse: dict | None = None,
    ) -> "Project":
        """Deferred; `template_reuse` reserves pins so non-matching devices skip."""
        self._field_device_defs.append((devices, reuse_terminals, template_reuse))
        return self

    def add_cable_runs(self, runs: list) -> "Project":
        """One cable page per `CableRun`, rendered after inter-device cables."""
        self._cable_runs.extend(runs)
        return self

    def route(
        self, *waypoints: "PinRef | Plc", net: "NetId | None" = None
    ) -> "Project":
        """Declare a multi-point signal for the owned Harness (resolved at build).

        Each waypoint is a concrete ``PinRef`` (device/terminal/PLC pin) or a
        ``Plc(signal_type, suffix)`` request to be allocated against the PLC rack.
        Additive to the legacy connection pipeline; see the Harness builder.

        Args:
            waypoints: Two or more route points (pins and/or Plc requests).
            net: Explicit net name; if omitted, the Harness synthesises one.

        Returns:
            self, for chaining.
        """
        self._route_decls.append((waypoints, net))
        return self

    def add_wires(self, wires: "Iterable[Wire]", /) -> "Project":
        """Ingest pre-built two-point ``Wire``s (e.g. explicit named-net wiring).

        Args:
            wires: Wires to add to the project's harness output at build.

        Returns:
            self, for chaining.
        """
        self._added_wires.extend(wires)
        return self

    # ------------------------------------------------------------------
    # Query properties
    # ------------------------------------------------------------------

    @property
    def device_registry(self) -> dict:
        """Merged device registry from all built circuits."""
        merged: dict = {}
        for result in self._results.values():
            merged.update(result.device_registry)
        return merged

    @property
    def bridge_groups(self) -> dict:
        """Merged bridge groups from all built circuits."""
        merged: dict = {}
        for result in self._results.values():
            for key, groups in result.bridge_groups.items():
                merged.setdefault(key, []).extend(groups)
        return merged

    @property
    def wire_connections(self) -> dict[str, list]:
        """Wire connections grouped by circuit name."""
        return {
            key: result.wire_connections
            for key, result in self._results.items()
            if result.wire_connections
        }

    @property
    def resolved_connections(self) -> list:
        """All resolved external connections (available after build_circuits())."""
        return list(self._external_connections)

    # ------------------------------------------------------------------
    # Page management
    # ------------------------------------------------------------------

    def page(self, title: str, circuit_key: str | list[str]) -> None:
        """A list of keys merges multiple circuits onto a single page."""
        if isinstance(circuit_key, list):
            self._pages.append(
                _PageDef(page_type="schematic", title=title, circuit_keys=circuit_key)
            )
        else:
            self._pages.append(
                _PageDef(page_type="schematic", title=title, circuit_key=circuit_key)
            )

    def front_page(self, md_path: str, notice: str | None = None) -> None:
        """Add a front page rendered from a Markdown file."""
        self._pages.append(_PageDef(page_type="front", md_path=md_path, notice=notice))

    def terminal_report(self) -> None:
        """Auto-generated page of registered terminals with bridge/connection info."""
        self._pages.append(_PageDef(page_type="terminal_report"))

    def plc_report(self, csv_path: str = "") -> "Project":
        """Empty `csv_path` auto-generates from `plc_rack()` (must be registered)."""
        self._pages.append(_PageDef(page_type="plc_report", csv_path=csv_path))
        return self

    def custom_page(self, title: str, typst_content: str) -> None:
        """Add a page with raw Typst markup content."""
        self._pages.append(
            _PageDef(page_type="custom", title=title, typst_content=typst_content)
        )

    def bom_report(self) -> "Project":
        """Add an auto-generated Bill of Materials page."""
        self._pages.append(_PageDef(page_type="bom_report"))
        return self

    def cable_pages(
        self,
        cable_prefix: str = "A-W",
        cable_start: int = 1,
        pins_last: tuple[str, ...] = ("PE",),
        temp_dir: str = "temp",
        *,
        toc: bool = True,
    ) -> "Project":
        """Render each cable to SVG via WireViz; needs `build_circuits()` first.

        `pins_last` reorders the named pins to the end of each cable.
        """
        from schematika.cable import (
            build_cable_drawings,
            cable_run_to_drawing,
            render_cable_svg,
        )

        # Collect all field devices
        all_devices: list = []
        for devices, _reuse, _template_reuse in self._field_device_defs:
            all_devices.extend(devices)

        # Build cable drawings from field device data
        drawings = build_cable_drawings(
            self._external_connections,
            all_devices,
            cable_prefix=cable_prefix,
            cable_start=cable_start,
            pins_last=pins_last,
            sort_integer_pins=self._pin_sort.sort_integers,
            sort_alphabetic_pins=self._pin_sort.sort_alphabetic,
        )

        # Append Wire-based cable runs (additive path), continuing the numbering.
        cable_number = cable_start + len(drawings)
        for run in self._cable_runs:
            drawings.append(
                cable_run_to_drawing(
                    run, f"{cable_prefix}{cable_number:03d}", self._pin_sort
                )
            )
            cable_number += 1

        # Render each cable to SVG file
        cable_dir = Path(temp_dir) / "cables"
        cable_dir.mkdir(parents=True, exist_ok=True)
        cable_entries: list[tuple[str, str, str, str]] = []
        for drawing in drawings:
            svg_content = render_cable_svg(drawing)
            svg_path = cable_dir / f"{drawing.cable.designator}.svg"
            with svg_path.open("w", encoding="utf-8") as f:
                f.write(svg_content)
            length_str = ""
            if drawing.cable.length:
                length_str = f"{drawing.cable.length:g} m"
            cable_entries.append(
                (str(svg_path), drawing.cable.designator, drawing.title, length_str)
            )

        # Add TOC page + cable pages (TOC skipped when toc=False)
        if toc:
            toc_entries = [
                (d.cable.designator, d.from_designator, ", ".join(d.to_designators))
                for d in drawings
            ]
            self._pages.append(
                _PageDef(
                    page_type="cable_toc",
                    title="Table of Contents",
                    cable_toc_entries=toc_entries,
                )
            )
        self._pages.append(
            _PageDef(
                page_type="cable",
                title="Cable Drawings",
                cable_entries=cable_entries,
            )
        )
        return self

    def export_wire_labels(
        self, path: str, titles: dict[str, str] | None = None
    ) -> "Project":
        """Register a wire label CSV export. Written during build()."""
        self._wire_label_export = (path, titles)
        return self

    def export_taglist(self, path: str) -> "Project":
        """Register a taglist CSV export. Written during build()."""
        self._taglist_export = path
        return self

    def export_bom_excel(self, path: str) -> "Project":
        """Register a BOM Excel export. Written during build()."""
        self._bom_excel_export = path
        return self

    def export_bom_csv(self, path: str) -> "Project":
        """Register a BOM CSV export. Written during build()."""
        self._bom_csv_export = path
        return self

    # ------------------------------------------------------------------
    # Build pipeline
    # ------------------------------------------------------------------

    def build_circuits(self) -> None:
        """Build all deferred circuits and resolve field devices."""
        self._build_all_circuits()
        self._resolve_field_devices()
        self._resolve_routes()

    def build(
        self,
        output: str,
        temp_dir: str = "temp",
        *,
        keep_temp: bool = False,
        datetime_stamp: bool = True,
    ) -> None:
        """Build circuits, render SVGs/CSVs, compile Typst -> PDF.

        `temp_dir` is removed after compile unless `keep_temp=True`.
        """
        from schematika.rendering.typst.compiler import (
            TypstCompiler,
            TypstCompilerConfig,
        )

        Path(temp_dir).mkdir(parents=True, exist_ok=True)

        # 1. Build all circuits
        self._build_all_circuits()
        self._resolve_field_devices()
        self._resolve_routes()

        # 2. Generate SVGs and terminal CSVs
        svg_paths = {}
        csv_paths = {}

        for key, result in self._results.items():
            svg_path = str(Path(temp_dir) / f"{key}.svg")
            _render_with_optional_pcb_viewbox(
                result.circuit, svg_path, self._pcb_page_viewboxes.get(key)
            )
            svg_paths[key] = svg_path

            if result.used_terminals:
                csv_path = str(Path(temp_dir) / f"{key}_terminals.csv")
                export_terminal_list(
                    csv_path, result.used_terminals, self._terminal_descriptions
                )
                csv_paths[key] = csv_path

        # 2b. Render P&ID SVGs
        pid_svg_paths = self._render_pid_svgs(temp_dir)

        self._render_multi_circuit_pages(svg_paths, csv_paths, temp_dir)
        self._export_wire_labels()
        self._export_taglist()
        self._export_bom_excel()
        self._export_bom_csv()

        # 3. Generate system terminal CSV with bridge info
        system_csv_path = self._emit_system_csv(temp_dir)

        # 3.5. Auto-generate PLC connections CSV if rack is configured
        plc_csv_path = ""
        if self._plc_rack is not None:
            plc_csv_path = str(Path(temp_dir) / "plc_connections.csv")
            self._generate_plc_csv(plc_csv_path)

        # 4. Assemble Typst document
        # Use CWD as root so all relative paths (SVGs, CSVs) resolve correctly
        root_dir = str(Path.cwd())
        config = TypstCompilerConfig(
            drawing_name=self.title,
            drawing_number=self.drawing_number,
            author=self.author,
            project=self.project,
            revision=self.revision,
            logo_path=str(Path(self.logo).resolve()) if self.logo else None,
            font_family=self.font,
            root_dir=root_dir,
            temp_dir=os.path.relpath(str(Path(temp_dir).resolve()), root_dir),
            datetime_stamp=datetime_stamp,
        )
        compiler = TypstCompiler(config)

        # Add pages
        for page_def in self._pages:
            self._add_page_to_compiler(
                compiler,
                page_def,
                svg_paths,
                csv_paths,
                system_csv_path,
                plc_csv_path,
                pid_svg_paths=pid_svg_paths,
            )

        # 5. Compile
        compiler.compile(output)

        # Cleanup
        if not keep_temp:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # ------------------------------------------------------------------
    # Build SVGs only (no PDF, no typst dependency)
    # ------------------------------------------------------------------

    def build_svgs(self, output_dir: str = "output") -> None:
        """Build circuits and export SVGs (no PDF; no `typst` dependency)."""
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        self._build_all_circuits()

        svg_paths: dict[str, str] = {}
        csv_paths: dict[str, str] = {}

        for key, result in self._results.items():
            svg_path = str(Path(output_dir) / f"{key}.svg")
            _render_with_optional_pcb_viewbox(
                result.circuit, svg_path, self._pcb_page_viewboxes.get(key)
            )
            svg_paths[key] = svg_path

            if result.used_terminals:
                csv_path = str(Path(output_dir) / f"{key}_terminals.csv")
                export_terminal_list(
                    csv_path, result.used_terminals, self._terminal_descriptions
                )
                csv_paths[key] = csv_path

        # Render P&ID SVGs
        self._render_pid_svgs(output_dir)

        self._render_multi_circuit_pages(svg_paths, csv_paths, output_dir)
        self._export_wire_labels()
        self._export_taglist()

        # System terminal CSV
        self._emit_system_csv(output_dir)

    # ------------------------------------------------------------------
    # Separate output methods
    # ------------------------------------------------------------------

    def render_svgs(self, output_dir: str) -> None:
        """Renders results already in `_results`; does NOT build deferred circuits."""
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        for key, result in self._results.items():
            svg_path = str(Path(output_dir) / f"{key}.svg")
            _render_with_optional_pcb_viewbox(
                result.circuit, svg_path, self._pcb_page_viewboxes.get(key)
            )

            if result.used_terminals:
                csv_path = str(Path(output_dir) / f"{key}_terminals.csv")
                export_terminal_list(
                    csv_path, result.used_terminals, self._terminal_descriptions
                )

        if self._pid_defs:
            self._render_pid_svgs(output_dir)

    def export_csvs(self, output_dir: str) -> None:
        """Writes the system terminal CSV; also PLC CSV if `plc_rack()` is set."""
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # System terminal CSV
        self._emit_system_csv(output_dir)

        # PLC connections CSV
        if self._plc_rack is not None:
            plc_csv_path = str(Path(output_dir) / "plc_connections.csv")
            self._generate_plc_csv(plc_csv_path)

    def compile_pdf(
        self,
        output: str,
        temp_dir: str = "temp",
        *,
        keep_temp: bool = False,
        datetime_stamp: bool = True,
    ) -> None:
        """Like `build()` but operates on results already in `_results`."""
        from schematika.rendering.typst.compiler import (
            TypstCompiler as _TypstCompiler,
        )
        from schematika.rendering.typst.compiler import (
            TypstCompilerConfig,
        )

        Path(temp_dir).mkdir(parents=True, exist_ok=True)

        # Render SVGs and per-circuit terminal CSVs
        svg_paths: dict[str, str] = {}
        csv_paths: dict[str, str] = {}

        for key, result in self._results.items():
            svg_path = str(Path(temp_dir) / f"{key}.svg")
            _render_with_optional_pcb_viewbox(
                result.circuit, svg_path, self._pcb_page_viewboxes.get(key)
            )
            svg_paths[key] = svg_path

            if result.used_terminals:
                csv_path = str(Path(temp_dir) / f"{key}_terminals.csv")
                export_terminal_list(
                    csv_path, result.used_terminals, self._terminal_descriptions
                )
                csv_paths[key] = csv_path

        # Build any P&ID diagrams not yet built (e.g. added after build_circuits())
        for pdef in self._pid_defs:
            if pdef.key not in self._pid_results:
                self._build_all_pids()
                break

        # Render P&ID SVGs
        pid_svg_paths = self._render_pid_svgs(temp_dir)

        self._render_multi_circuit_pages(svg_paths, csv_paths, temp_dir)
        self._export_wire_labels()
        self._export_taglist()

        # System terminal CSV with bridge info
        system_csv_path = self._emit_system_csv(temp_dir)

        # PLC connections CSV
        plc_csv_path = ""
        if self._plc_rack is not None:
            plc_csv_path = str(Path(temp_dir) / "plc_connections.csv")
            self._generate_plc_csv(plc_csv_path)

        # Assemble Typst document
        root_dir = str(Path.cwd())
        config = TypstCompilerConfig(
            drawing_name=self.title,
            drawing_number=self.drawing_number,
            author=self.author,
            project=self.project,
            revision=self.revision,
            logo_path=str(Path(self.logo).resolve()) if self.logo else None,
            font_family=self.font,
            root_dir=root_dir,
            temp_dir=os.path.relpath(str(Path(temp_dir).resolve()), root_dir),
            datetime_stamp=datetime_stamp,
        )
        compiler = _TypstCompiler(config)

        for page_def in self._pages:
            self._add_page_to_compiler(
                compiler,
                page_def,
                svg_paths,
                csv_paths,
                system_csv_path,
                plc_csv_path,
                pid_svg_paths=pid_svg_paths,
            )

        compiler.compile(output)

        if not keep_temp:
            shutil.rmtree(temp_dir, ignore_errors=True)

    # ------------------------------------------------------------------
    # Internal: field device resolution
    # ------------------------------------------------------------------

    def _resolve_harness(self) -> HarnessBuildResult:
        """Resolve buffered route() declarations + add_wires() into one result.

        Builds a Harness bound to the project's PLC rack (empty if none set),
        replays the buffered routes, and folds directly-added wires into the
        frozen result. Does not mutate the project; builds a fresh Harness from
        the buffered declarations each call. Build-time entry point of the Wire
        pipeline.
        """
        harness = Harness(rack=self._plc_rack if self._plc_rack is not None else [])
        for waypoints, net in self._route_decls:
            harness.route(*waypoints, net=net)
        result = harness.build()
        if self._added_wires:
            result = HarnessBuildResult(
                wires=(*result.wires, *self._added_wires),
                plc_assignments=result.plc_assignments,
            )
        return result

    def _resolve_routes(self) -> None:
        """Resolve buffered route()/add_wires() into terminal-only rows.

        Additive Wire-pipeline replacement for ``internal_wiring()``: builds
        the harness from the buffered declarations and appends the converted
        ConnectionRows to ``_terminal_only_connections``. No-op when no route()
        or add_wires() declarations were made.
        """
        if not self._route_decls and not self._added_wires:
            return
        result = self._resolve_harness()
        rows = _wires_to_terminal_rows(result)
        self._terminal_only_connections.extend(rows)
        self._route_terminal_rows.extend(rows)

    def _resolve_field_devices(self) -> None:
        """Resolve deferred field device registrations.

        Legacy tuple path (``_external_connections``) and the additive Harness
        wire path run from the SAME resolved reuse: the legacy tuples feed
        BOM/legacy CSV; the harness wires + correlated ``PlcAssignment``s feed
        the native terminal CSV (byte-identical) and the PLC report (R3).
        """
        if not self._field_device_defs:
            return

        from schematika.electrical.field_devices import generate_field_connections
        from schematika.electrical.terminal_emit import field_device_rows

        for devices, reuse_terminals, template_reuse in self._field_device_defs:
            resolved_reuse = self._resolve_terminal_reuse(reuse_terminals)
            resolved_template_reuse = self._resolve_template_reuse(template_reuse)

            connections = generate_field_connections(
                devices,
                reuse_terminals=resolved_reuse,
                template_reuse=resolved_template_reuse,
            )

            if self._plc_rack:
                from schematika.electrical.plc_resolver import resolve_plc_references

                connections = resolve_plc_references(connections, self._plc_rack)

            self._external_connections.extend(connections)
            self._field_device_rows.extend(connections)

            harness = Harness(rack=self._plc_rack if self._plc_rack is not None else [])
            harness.add_field_devices(
                devices,
                reuse_terminals=resolved_reuse,
                template_reuse=resolved_template_reuse,
            )
            result = harness.build()
            self._plc_assignments.extend(result.plc_assignments)
            self._field_device_wires.extend(result.wires)
            self._field_device_wire_rows.extend(
                field_device_rows(result.wires, result.plc_assignments)
            )

    def _resolve_terminal_reuse(self, reuse_terminals: dict | None) -> dict | None:
        """Resolve terminal reuse dict to built circuit results."""
        if not reuse_terminals:
            return None
        resolved: dict = {}
        for terminal, circuit_key in reuse_terminals.items():
            if circuit_key not in self._results:
                msg = (
                    f"field_devices() references circuit '{circuit_key}' "
                    f"for terminal reuse, but it hasn't been built yet."
                )
                raise CircuitValidationError(msg)
            resolved[str(terminal)] = self._results[circuit_key]
        return resolved

    def _resolve_template_reuse(self, template_reuse: dict | None) -> dict | None:
        """Resolve template reuse dict to built circuit results."""
        if not template_reuse:
            return None
        resolved: dict = {}
        for tmpl, terminal_map in template_reuse.items():
            resolved[tmpl] = {}
            for terminal, circuit_key in terminal_map.items():
                if circuit_key not in self._results:
                    msg = (
                        f"field_devices() template_reuse references "
                        f"circuit '{circuit_key}', but it hasn't "
                        f"been built yet."
                    )
                    raise CircuitValidationError(msg)
                resolved[tmpl][str(terminal)] = self._results[circuit_key]
        return resolved

    # ------------------------------------------------------------------
    # Internal: circuit building
    # ------------------------------------------------------------------

    def _build_all_circuits(self) -> None:
        """Build all registered circuits and P&ID diagrams in order."""
        self._results = {}
        for cdef in self._circuit_defs:
            result = self._build_one_circuit(cdef)
            self._results[cdef.key] = result
            self._state = result.state
        self._build_all_pids()

    def _build_all_pids(self) -> None:
        """Build all registered P&ID diagram definitions in order."""
        from schematika.pid.builder import PIDBuilder

        self._pid_results = {}
        for pdef in self._pid_defs:
            builder_or_factory = pdef.builder_or_factory
            if isinstance(builder_or_factory, PIDBuilder):
                result = builder_or_factory.build(state=self._state)
            elif callable(builder_or_factory):
                result = builder_or_factory(self._state)
            else:
                msg = (
                    f"add_pid('{pdef.key}'): builder_or_factory must be a "
                    f"PIDBuilder instance or a callable, got "
                    f"{type(builder_or_factory).__name__}"
                )
                raise TypeError(msg)
            self._pid_results[pdef.key] = result
            self._state = result.state

    def _render_pid_svgs(self, output_dir: str) -> dict[str, str]:
        """Returns a mapping of diagram key -> SVG file path."""
        from schematika.pid.diagram import render_pid

        pid_svg_paths: dict[str, str] = {}
        for key, result in self._pid_results.items():
            svg_path = str(Path(output_dir) / f"pid_{key}.svg")
            render_pid(result.diagram, svg_path)
            pid_svg_paths[key] = svg_path
        return pid_svg_paths

    def _build_one_circuit(self, cdef: _CircuitDef) -> BuildResult:
        """Build a single circuit definition."""
        # Resolve reuse_tags: map circuit key -> BuildResult
        resolved_reuse = None
        if cdef.reuse_tags:
            resolved_reuse = {}
            for prefix, source_key in cdef.reuse_tags.items():
                if source_key not in self._results:
                    msg = (
                        f"Circuit '{cdef.key}' references '{source_key}' via "
                        f"reuse_tags, but it hasn't been built yet. "
                        f"Register '{source_key}' before '{cdef.key}'."
                    )
                    raise CircuitValidationError(msg)
                resolved_reuse[prefix] = self._results[source_key]

        if cdef.factory == "descriptors":
            return self._build_descriptor_circuit(cdef, resolved_reuse)
        if cdef.factory == "custom":
            return self._build_custom_circuit(cdef)
        msg = (
            f"Unknown circuit factory '{cdef.factory}'. Use 'descriptors' or 'custom'."
        )
        raise CircuitValidationError(msg)

    def _build_descriptor_circuit(
        self, cdef: _CircuitDef, resolved_reuse: dict | None
    ) -> BuildResult:
        """Build a circuit from inline descriptors."""
        if cdef.components is None:
            msg = (
                f"Circuit '{cdef.key}' uses descriptor mode but has no "
                f"components defined"
            )
            raise CircuitValidationError(msg)
        return build_from_descriptors(
            self._state,
            cdef.components,
            options=DescriptorBuildOptions(
                x=cdef.params.get("x", 0.0),
                y=cdef.params.get("y", 0.0),
                spacing=cdef.params.get("spacing", 80.0),
                count=cdef.count,
                wire_labels=tuple(cdef.wire_labels) if cdef.wire_labels else None,
                reuse_tags=resolved_reuse,
                start_indices=cdef.start_indices,
                terminal_start_indices=cdef.terminal_start_indices,
            ),
        )

    def _build_custom_circuit(self, cdef: _CircuitDef) -> BuildResult:
        """Build a circuit via user-provided builder function."""
        if cdef.builder_fn is None:
            msg = f"Circuit '{cdef.key}' uses custom mode but has no builder_fn defined"
            raise CircuitValidationError(msg)
        result = cdef.builder_fn(self._state, **cdef.params)
        if isinstance(result, BuildResult):
            return result
        # Support frozen CircuitBuilder return
        if isinstance(result, CircuitBuilder) and result._frozen:
            return result._result
        # Support tuple return: (state, circuit, used_terminals)
        state, circuit, used_terminals = result
        return BuildResult(
            state=state,
            circuit=circuit,
            used_terminals=used_terminals,
        )

    def _render_multi_circuit_pages(
        self,
        svg_paths: dict[str, str],
        csv_paths: dict[str, str],
        output_dir: str,
    ) -> None:
        """Render merged SVGs for multi-circuit pages."""
        from schematika.electrical.builder import merge_build_results

        for page_def in self._pages:
            if page_def.circuit_keys:
                results_to_merge = [
                    self._results[k]
                    for k in page_def.circuit_keys
                    if k in self._results
                ]
                if results_to_merge:
                    merged = merge_build_results(results_to_merge)
                    merged_key = "_".join(page_def.circuit_keys)
                    svg_path = str(Path(output_dir) / f"{merged_key}.svg")
                    # Preserve the PCB page viewBox + dims (if any circuit on this
                    # page has one); otherwise render auto-fit.
                    page_dims = next(
                        (
                            self._pcb_page_viewboxes[k]
                            for k in page_def.circuit_keys
                            if k in self._pcb_page_viewboxes
                        ),
                        None,
                    )
                    _render_with_optional_pcb_viewbox(
                        merged.circuit, svg_path, page_dims
                    )
                    svg_paths[merged_key] = svg_path
                    if merged.used_terminals:
                        csv_path_m = str(
                            Path(output_dir) / f"{merged_key}_terminals.csv"
                        )
                        export_terminal_list(
                            csv_path_m,
                            merged.used_terminals,
                            self._terminal_descriptions,
                        )
                        csv_paths[merged_key] = csv_path_m
                    # Point page_def to merged key for compiler
                    page_def.circuit_key = merged_key

    # ------------------------------------------------------------------
    # Internal: wire label and taglist exports
    # ------------------------------------------------------------------

    def _export_wire_labels(self) -> None:
        if self._wire_label_export is None:
            return
        import csv as _csv

        path, titles = self._wire_label_export
        titles = titles or {}
        Path(path).resolve().parent.mkdir(parents=True, exist_ok=True)
        with Path(path).open("w", newline="") as f:
            writer = _csv.writer(f)
            for circuit_key, result in self._results.items():
                if not result.wire_connections:
                    continue
                title = titles.get(circuit_key, circuit_key)
                writer.writerow([title])
                for tag_a, pin_a, tag_b, pin_b in result.wire_connections:
                    writer.writerow([f"{tag_a}:{pin_a}"])
                    writer.writerow([f"{tag_b}:{pin_b}"])
                writer.writerow([])

    def _export_taglist(self) -> None:
        if self._taglist_export is None:
            return
        import csv as _csv

        from schematika.core.utils import natural_sort_key

        tags: set[str] = set()
        for result in self._results.values():
            tags.update(result.device_registry.keys())
        for tid in self._terminals:
            tags.add(tid)
        if self._plc_rack:
            for slot_name, _module in self._plc_rack:
                tags.add(slot_name)

        Path(self._taglist_export).resolve().parent.mkdir(parents=True, exist_ok=True)
        with Path(self._taglist_export).open("w", newline="") as f:
            writer = _csv.writer(f)
            writer.writerow(["Tag"])
            for tag in sorted(tags, key=natural_sort_key):
                writer.writerow([tag])

    def _export_bom_excel(self) -> None:
        if self._bom_excel_export is None:
            return
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Font,
            PatternFill,
        )

        rows = self._aggregate_bom()
        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        headers = ["Tags", "MPN", "Description", "Qty"]
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

        for row_idx, (tags, mpn, desc, qty) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=tags)
            ws.cell(row=row_idx, column=2, value=mpn)
            ws.cell(row=row_idx, column=3, value=desc)
            ws.cell(row=row_idx, column=4, value=qty).alignment = Alignment(
                horizontal="right"
            )

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 8

        path = self._bom_excel_export
        Path(path).resolve().parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    def _export_bom_csv(self) -> None:
        if self._bom_csv_export is None:
            return
        import csv

        rows = self._aggregate_bom()
        path = self._bom_csv_export
        Path(path).resolve().parent.mkdir(parents=True, exist_ok=True)
        with Path(path).open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Tags", "MPN", "Description", "Qty"])
            for tags, mpn, desc, qty in rows:
                writer.writerow([tags, mpn, desc, qty])

    # ------------------------------------------------------------------
    # Internal: system CSV generation
    # ------------------------------------------------------------------

    def _emit_system_csv(self, output_dir: str) -> str:
        """Dispatch to the native (opt-in) or legacy terminal-CSV path."""
        if self._native_terminal_emit:
            return self._generate_system_csv_native(output_dir)
        return self._generate_system_csv(output_dir)

    def _generate_system_csv(self, output_dir: str) -> str:
        """PLC-prefixed connections are filtered out (they go in the PLC report)."""
        from schematika.electrical.plc_resolver import PLC_PREFIX
        from schematika.electrical.system.connection_registry import TerminalRegistry

        csv_path = str(Path(output_dir) / "system_terminals.csv")
        registry = get_registry(self._state)
        filtered = tuple(
            c for c in registry.connections if not c.terminal_tag.startswith(PLC_PREFIX)
        )
        registry = TerminalRegistry(connections=filtered)
        export_registry_to_csv(registry, csv_path, state=self._state)

        # Bridge defs from Terminal objects
        bridge_defs: dict = {}
        prefix_bridge_tags: set[str] = set()
        for tid, t in self._terminals.items():
            if t.bridge and not t.reference:
                if t.bridge == BridgeMode.PER_PREFIX:
                    prefix_bridge_tags.add(tid)
                else:
                    bridge_defs[tid] = t.bridge

        # Bridge groups from circuit results
        for result in self._results.values():
            for key, groups in result.bridge_groups.items():
                bridge_defs.setdefault(key, []).extend(groups)

        finalize_terminal_csv(
            csv_path,
            bridge_defs=bridge_defs or None,
            prefix_bridge_tags=prefix_bridge_tags or None,
            external_connections=(
                self._external_connections + self._terminal_only_connections
            )
            or None,
        )
        return csv_path

    def _generate_system_csv_native(self, output_dir: str) -> str:
        """Native C3a terminal CSV; byte-identical to _generate_system_csv."""
        from collections import Counter, defaultdict

        from schematika.electrical.plc_resolver import PLC_PREFIX
        from schematika.electrical.system.connection_registry import (
            TerminalRegistry,
            _build_all_pin_keys,
        )
        from schematika.electrical.terminal_emit import (
            panel_terminal_emit,
            terminal_csv_rows,
        )

        csv_path = str(Path(output_dir) / "system_terminals.csv")
        registry = get_registry(self._state)
        filtered = tuple(
            c for c in registry.connections if not c.terminal_tag.startswith(PLC_PREFIX)
        )
        registry = TerminalRegistry(connections=filtered)

        grouped: dict = defaultdict(lambda: {"top": [], "bottom": []})
        for conn in registry.connections:
            grouped[(conn.terminal_tag, conn.terminal_pin)][conn.side].append(conn)
        allocated_pin_keys = tuple(_build_all_pin_keys(grouped, self._state))

        wires, sidecar = panel_terminal_emit(
            registry,
            self._terminals,
            allocated_pin_keys=allocated_pin_keys,
            bridge_groups=self.bridge_groups,
        )
        # Route/add_wires rows are re-emitted natively (first-waypoint-anchored)
        # as route_wires, so drop their buggy tuples (wrong-end anchored) from
        # the verbatim external set.
        # Multiset subtraction preserves count if a hand tuple equals a route tuple.
        remaining = Counter(self._route_terminal_rows)
        terminal_only: list = []
        for r in self._terminal_only_connections:
            if remaining.get(r, 0) > 0:
                remaining[r] -= 1
            else:
                terminal_only.append(r)
        # Field-device tuples are re-emitted natively from harness wires below;
        # subtract them from the verbatim external set (same multiset de-dup).
        fd_remaining = Counter(self._field_device_rows)
        external_only: list = []
        for r in self._external_connections:
            if fd_remaining.get(r, 0) > 0:
                fd_remaining[r] -= 1
            else:
                external_only.append(r)
        external_rows = external_only + terminal_only + self._field_device_wire_rows
        route_wires: tuple[tuple[Wire, TerminalWireFact], ...] = ()
        if self._route_decls or self._added_wires:
            route_wires = tuple(_wires_to_terminal_facts(self._resolve_harness()))
        terminal_csv_rows(wires, sidecar, external_rows, csv_path, route_wires)
        return csv_path

    # ------------------------------------------------------------------
    # Internal: BOM aggregation
    # ------------------------------------------------------------------

    def _count_terminal_pins(self) -> dict[str, int]:
        """Count unique pins per terminal from all connection sources."""
        pin_sets: dict[str, set] = {}
        for conn in get_registry(self._state).connections:
            pin_sets.setdefault(conn.terminal_tag, set()).add(conn.terminal_pin)
        for row in self._external_connections:
            tag, pin = str(row[2]), row[3]
            if tag and pin:
                pin_sets.setdefault(tag, set()).add(pin)
        for row in self._terminal_only_connections:
            tag, pin = str(row[2]), row[3]
            if tag and pin:
                pin_sets.setdefault(tag, set()).add(pin)
        return {k: len(v) for k, v in pin_sets.items()}

    def _aggregate_bom(self) -> list[tuple[str, str, str, int]]:
        """Aggregate BOM from device registries, terminals, and PLC modules."""
        from schematika.core.utils import natural_sort_key

        terminal_pin_counts = self._count_terminal_pins()

        # Devices
        device_groups: dict[tuple[str, str], list[str]] = {}
        for result in self._results.values():
            for tag, device in result.device_registry.items():
                key = (device.mpn, device.description)
                device_groups.setdefault(key, []).append(tag)

        rows: list[tuple[str, str, str, int]] = []
        for (mpn, desc), tags in sorted(device_groups.items()):
            unique_tags = sorted(set(tags), key=natural_sort_key)
            rows.append(("/".join(unique_tags), mpn, desc, len(unique_tags)))

        # Terminals
        terminal_groups: dict[tuple[str, str], list[tuple[str, int]]] = {}
        for tid, t in self._terminals.items():
            if not t.reference and t.mpn:
                key = (t.mpn, t.description)
                pin_count = terminal_pin_counts.get(tid, 0)
                terminal_groups.setdefault(key, []).append((tid, pin_count))

        for (mpn, desc), entries in sorted(terminal_groups.items()):
            tags = sorted([e[0] for e in entries], key=natural_sort_key)
            total_pins = sum(e[1] for e in entries)
            rows.append(("/".join(tags), mpn, desc, total_pins))

        # PLC modules
        if self._plc_rack:
            plc_groups: dict[str, list[str]] = {}
            plc_desc: dict[str, str] = {}
            for slot_name, module in self._plc_rack:
                plc_groups.setdefault(module.mpn, []).append(slot_name)
                plc_desc[module.mpn] = (
                    f"PLC {module.signal_type} module ({module.channels}ch)"
                )
            for mpn, slots in sorted(plc_groups.items()):
                sorted_slots = sorted(slots, key=natural_sort_key)
                rows.append(
                    ("/".join(sorted_slots), mpn, plc_desc[mpn], len(sorted_slots))
                )

        return rows

    def _generate_bom_typst(self, bom_rows: list[tuple[str, str, str, int]]) -> str:
        """Generate Typst markup for BOM table."""
        lines = [
            "#place(bottom + center, dy: -title_offset)[",
            '  #text(size: 18pt, weight: "bold")[Bill of Materials]',
            "]",
            "#pad(left: 25mm, right: 25mm, top: 40mm, bottom: 40mm)[",
            "  #columns(2, gutter: 30em)[",
            "    #block(breakable: true)[",
            "      #table(",
            "        columns: (4.5cm, 3.5cm, 1fr, 1cm),",
            "        align: (left, left, left, right),",
            "        fill: (x, y) => if y == 0 { gray.lighten(85%) } else { none },",
            "        inset: 4pt,",
            "        stroke: 0.25pt + gray,",
            "        table.header(",
            '          text(size: 9pt, weight: "bold")[Tags],',
            '          text(size: 9pt, weight: "bold")[MPN],',
            '          text(size: 9pt, weight: "bold")[Description],',
            '          text(size: 9pt, weight: "bold")[Qty],',
            "        ),",
        ]
        for tags, mpn, desc, qty in bom_rows:
            tags_esc = tags.replace("#", "\\#")
            mpn_esc = mpn.replace("#", "\\#")
            desc_esc = desc.replace("#", "\\#")
            lines.append(
                f"        text(size: 9pt)[{tags_esc}], "
                f"text(size: 9pt)[{mpn_esc}], "
                f"text(size: 9pt)[{desc_esc}], "
                f"text(size: 9pt)[{qty}],"
            )
        lines.append("      )")
        lines.append("    ]")
        lines.append("  ]")
        lines.append("]")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Internal: page compilation
    # ------------------------------------------------------------------

    def _add_schematic_page(
        self,
        compiler: "TypstCompiler",
        page_def: _PageDef,
        svg_paths: dict[str, str],
        csv_paths: dict[str, str],
        pid_svg_paths: dict[str, str] | None,
    ) -> None:
        """Resolve SVG/CSV for a schematic or P&ID page and register it."""
        svg_path, csv_path = _resolve_svg_for_page(
            page_def.page_type,
            page_def.circuit_key,
            svg_paths,
            csv_paths,
            pid_svg_paths,
        )
        if svg_path:
            compiler.add_schematic_page(page_def.title, svg_path, csv_path)

    def _add_page_to_compiler(
        self,
        compiler: "TypstCompiler",
        page_def: _PageDef,
        svg_paths: dict[str, str],
        csv_paths: dict[str, str],
        system_csv_path: str,
        plc_csv_path: str = "",
        pid_svg_paths: dict[str, str] | None = None,
    ) -> None:
        """Add a page definition to the TypstCompiler."""
        match page_def.page_type:
            case "schematic" | "pid":
                self._add_schematic_page(
                    compiler, page_def, svg_paths, csv_paths, pid_svg_paths
                )
            case "front":
                compiler.add_front_page(page_def.md_path, notice=page_def.notice)
            case "terminal_report":
                titles = {
                    str(t): t.title for t in self._terminals.values() if not t.reference
                }
                compiler.add_terminal_report(system_csv_path, titles)
            case "plc_report":
                csv_path = page_def.csv_path or plc_csv_path
                if csv_path:
                    compiler.add_plc_report(csv_path)
            case "custom":
                compiler.add_custom_page(page_def.title, page_def.typst_content)
            case "bom_report":
                bom_rows = self._aggregate_bom()
                typst_content = self._generate_bom_typst(bom_rows)
                compiler.add_custom_page("Bill of Materials", typst_content)
            case "cable" if page_def.cable_entries:
                compiler.add_cable_pages(page_def.cable_entries)
            case "cable_toc" if page_def.cable_toc_entries:
                compiler.add_cable_toc(page_def.cable_toc_entries)

    # ------------------------------------------------------------------
    # Internal: PLC CSV generation
    # ------------------------------------------------------------------

    def _generate_plc_csv(self, csv_path: str) -> None:
        """Generate PLC connections CSV from registry and external connections.

        Opt-in native path (``use_native_plc_report``) rebuilds the report from
        the buffered field-device ``HarnessBuildResult``; it is byte-identical to
        the legacy 3-source pipeline ONLY when no registry-logged PLC connections
        exist, so it falls back to legacy when there is no field-device result.
        """
        if self._native_plc_report and self._field_device_wires:
            self._generate_plc_csv_native(csv_path)
            return

        import csv as _csv

        from schematika.electrical.plc_resolver import (
            extract_plc_connections_from_registry,
            generate_plc_report_rows,
            resolve_plc_references,
        )

        # _generate_plc_csv is only called when _plc_rack is not None
        rack = self._plc_rack
        assert rack is not None  # noqa: S101 — _generate_plc_csv is only called when _plc_rack is not None

        # Resolve external connections if any
        external = list(self._external_connections)
        if external:
            external = resolve_plc_references(external, rack)

        # Extract registry connections
        registry_connections: list[ConnectionRow] = (
            extract_plc_connections_from_registry(self._state, rack, external)
        )

        # Merge and generate rows
        all_connections = external + registry_connections
        rows = generate_plc_report_rows(all_connections, rack)

        with Path(csv_path).open("w", newline="") as f:
            writer = _csv.writer(f)
            writer.writerow(
                ["Module", "MPN", "PLC Pin", "Component", "Pin", "Terminal"]
            )
            writer.writerows(rows)

    def _generate_plc_csv_native(self, csv_path: str) -> None:
        """Native PLC report via plc_csv_rows; byte-identical to _generate_plc_csv."""
        import csv as _csv

        from schematika.electrical.plc_report import plc_csv_rows

        rack = self._plc_rack
        assert rack is not None  # noqa: S101 — only called when _plc_rack is not None

        result = HarnessBuildResult(
            wires=tuple(self._field_device_wires),
            plc_assignments=tuple(self._plc_assignments),
        )
        rows = plc_csv_rows(result, rack)

        with Path(csv_path).open("w", newline="") as f:
            writer = _csv.writer(f)
            writer.writerow(
                ["Module", "MPN", "PLC Pin", "Component", "Pin", "Terminal"]
            )
            writer.writerows(rows)
